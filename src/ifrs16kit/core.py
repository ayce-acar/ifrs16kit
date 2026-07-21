#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════════════════
 ifrs16kit — Interactive IFRS 16 Re-performance Tool  (LIQUID edition)
═══════════════════════════════════════════════════════════════════════════
 Standalone script. Works in VS Code, Spyder, PyCharm, or plain terminal.

 Requires:  pip install openpyxl

 What it does
 ------------
 1. Asks you about the lease (country, payment timing, frequency)
 2. Generates IFRS16_Input.xlsx — a LIQUID input template
    (dark-red cells = your inputs, everything else auto-calculates
     with live Excel formulas)
 3. Waits for you to fill it in
 4. Reads the template back, validates the inputs
 5. Generates IFRS16_Calculation.xlsx — a fully formula-driven
    10-sheet re-performance workbook:
      Inputs · Initial Measurement · Lease Schedule · Annual Summary
      · Journals · Tax_Reconciliation · Findings · Audit_Procedures
      · PBC_List · Setup
    Every derived figure in the workbook is a live Excel formula,
    so the file stays "liquid": change any input and the whole
    model recalculates.
 6. Prints an independent Python cross-check of the key figures
    (lease liability, ROU asset, total interest) so the auditor can
    verify the Excel output against a second, independent engine.

 Usage
 -----
   ifrs16kit                            → guided interactive flow
   ifrs16kit --demo                     → skip questions, run the CALISMA
                                          golden case end-to-end instantly

 Golden benchmark (CALISMA.xlsx)
 -------------------------------
   Monthly advance · 24 periods · 6% IBR · 650/month · 600 IDC
   → Liability 14,761.09 · ROU 15,361.09 · Interest 838.91

 Author: Ayce Acar — MSc Business Analytics, TCD
═══════════════════════════════════════════════════════════════════════════
"""

import os
import sys
from dataclasses import dataclass
from datetime import date, datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required.  Run:  pip install openpyxl")


# ═══════════════════════════════════════════════════════════════════════
#  CONFIG — countries, frequencies, capacity, styles
# ═══════════════════════════════════════════════════════════════════════

# Mirrors the Setup sheet of the calculation workbook exactly.
COUNTRIES = {
    "Türkiye":        {"cur": "TRY", "sym": "₺",   "framework": "TFRS (KGK)",
                       "std": "TFRS 16", "isa540": "BDS 540",
                       "isa500": "BDS 500", "isa230": "BDS 230", "tax": 0.25},
    "Ireland":        {"cur": "EUR", "sym": "€",   "framework": "IFRS (EU-adopted)",
                       "std": "IFRS 16", "isa540": "ISA (Ireland) 540",
                       "isa500": "ISA (Ireland) 500", "isa230": "ISA (Ireland) 230",
                       "tax": 0.125},
    "United Kingdom": {"cur": "GBP", "sym": "£",   "framework": "UK-adopted IFRS",
                       "std": "IFRS 16", "isa540": "ISA (UK) 540",
                       "isa500": "ISA (UK) 500", "isa230": "ISA (UK) 230",
                       "tax": 0.25},
    "UAE":            {"cur": "AED", "sym": "د.إ", "framework": "IFRS",
                       "std": "IFRS 16", "isa540": "ISA 540",
                       "isa500": "ISA 500", "isa230": "ISA 230", "tax": 0.09},
    "Australia":      {"cur": "AUD", "sym": "$",   "framework": "AASB (IFRS-equivalent)",
                       "std": "AASB 16", "isa540": "ASA 540",
                       "isa500": "ASA 500", "isa230": "ASA 230", "tax": 0.30},
}

FREQUENCIES = {12: "Monthly", 4: "Quarterly", 2: "Semi-annual", 1: "Annual"}
MAX_PERIODS = 120                      # template row capacity

TEMPLATE_NAME = "IFRS16_Input.xlsx"
OUTPUT_NAME   = "IFRS16_Calculation.xlsx"

# ── Style constants (LIQUID model conventions) ──────────────────────────
NAVY_FILL   = PatternFill("solid", start_color="002060")   # section headers
RED_FILL    = PatternFill("solid", start_color="C0504D")   # dark-red inputs
GREEN_FILL  = PatternFill("solid", start_color="E2EFDA")   # key outputs / checks

F_TITLE  = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
F_HDR    = Font(name="Times New Roman", size=12, bold=True, color="FFFFFF")
F_SECT   = Font(name="Times New Roman", size=12, bold=True)
F_LABEL  = Font(name="Times New Roman", size=11)
F_BODY   = Font(name="Times New Roman", size=11)
F_BOLD   = Font(name="Times New Roman", size=12, bold=True)
F_INPUT  = Font(name="Times New Roman", size=11, bold=True)          # dark-red cells
F_NOTE   = Font(name="Times New Roman", size=10, color="595959")     # grey notes
F_LINK   = Font(name="Times New Roman", size=11, color="008000")     # green = link
F_CHECK  = Font(name="Times New Roman", size=11, color="008000")     # green checks

MEDIUM = Side(style="medium", color="000000")
BOX_M  = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FMT_MONEY   = "#,##0.00"
FMT_MONEY0  = "#,##0"
FMT_DATE    = "dd\\-mmm\\-yyyy"
FMT_FACTOR  = "0.000000"
FMT_PCT2    = "0.00%"
FMT_PCT4    = "0.0000%"


# ═══════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════

@dataclass
class LeaseInputs:
    entity: str = "Example Co. Limited"
    asset: str = "Company car (illustrative)"
    country: str = "Ireland"
    commencement: date = date(2025, 1, 1)
    term_years: int = 2
    term_extra_months: int = 0
    freq: int = 12                      # payments per year
    payment: float = 650.0
    is_advance: bool = True
    ibr_annual: float = 0.06
    idc: float = 600.0
    prepaid: float = 0.0
    incentives: float = 0.0
    restoration: float = 0.0
    method: str = "Straight-line"       # or "Reducing balance"
    useful_life_months: int = 60
    rb_rate: float = 0.30               # reducing-balance annual rate

    @property
    def term_months(self) -> int:
        return self.term_years * 12 + self.term_extra_months

    @property
    def months_per_period(self) -> float:
        return 12 / self.freq

    @property
    def num_periods(self) -> float:
        return self.term_months / self.months_per_period

    @property
    def timing_flag(self) -> int:
        return 1 if self.is_advance else 0

    @property
    def periodic_rate(self) -> float:
        return (1 + self.ibr_annual) ** (1 / self.freq) - 1


# ═══════════════════════════════════════════════════════════════════════
#  PYTHON CROSS-CHECK ENGINE
#  (mirrors the Excel formulas 1:1 — independent verification of the
#   generated workbook, IFRS 16 paras 24, 26, 36; ISA 500 A21)
# ═══════════════════════════════════════════════════════════════════════

def cross_check(inp: LeaseInputs) -> dict:
    r    = inp.periodic_rate
    n    = int(inp.num_periods)
    flag = inp.timing_flag

    # Lease liability = PV of payments (para 26); factor 1/(1+r)^(p-flag)
    liability = sum(inp.payment / (1 + r) ** (p - flag) for p in range(1, n + 1))

    # ROU asset (para 24)
    rou = liability + inp.idc + inp.prepaid - inp.incentives + inp.restoration

    # Schedule (para 36): interest = (opening − flag·payment)·r
    depr_periods = min(n, int(inp.useful_life_months / inp.months_per_period))
    rb_periodic  = 1 - (1 - inp.rb_rate) ** (1 / inp.freq)

    o_liab, o_rou = liability, rou
    tot_int = tot_dep = 0.0
    for p in range(1, n + 1):
        interest = (o_liab - flag * inp.payment) * r
        c_liab   = o_liab - inp.payment + interest
        if p > depr_periods:
            dep = 0.0
        elif inp.method == "Straight-line":
            dep = rou / depr_periods
        else:                                    # Reducing balance
            dep = o_rou if p == depr_periods else o_rou * rb_periodic
        tot_int += interest
        tot_dep += dep
        o_liab, o_rou = c_liab, o_rou - dep

    return {"periodic_rate": r, "liability": liability, "rou": rou,
            "total_interest": tot_int, "total_depreciation": tot_dep,
            "closing_liability": o_liab, "closing_rou": o_rou,
            "n": n, "total_payments": inp.payment * n}


# ═══════════════════════════════════════════════════════════════════════
#  STEP 1 — INTERACTIVE QUESTIONS
# ═══════════════════════════════════════════════════════════════════════

def ask_choice(prompt: str, labels: list, default_idx: int) -> int:
    print(f"\n{prompt}")
    for i, lab in enumerate(labels, 1):
        mark = "  ← default" if (i - 1) == default_idx else ""
        print(f"   {i}.  {lab}{mark}")
    while True:
        raw = input(f"   Choose 1-{len(labels)} [Enter = default]: ").strip()
        if raw == "":
            return default_idx
        if raw.isdigit() and 1 <= int(raw) <= len(labels):
            return int(raw) - 1
        print("   Invalid — try again.")


def ask_yes_no(prompt: str, default: bool) -> bool:
    d = "Y/n" if default else "y/N"
    raw = input(f"\n{prompt}  [{d}]: ").strip().lower()
    if raw == "":
        return default
    return raw.startswith("y")


def interview() -> dict:
    print("═" * 70)
    print("  ifrs16kit — IFRS 16 Auditor Re-performance  (LIQUID model)")
    print("═" * 70)
    print("""
  This tool re-performs an IFRS 16 lease calculation from first
  principles — independently of the client — using the effective
  interest method (IFRS 16, para 36).

  Flow:  answer 3 questions  →  fill in the Excel input template
         →  receive the full LIQUID calculation workbook.
""")
    countries = list(COUNTRIES)
    ci = ask_choice("1)  Which country / jurisdiction is the lease in?",
                    [f"{c}  ({COUNTRIES[c]['cur']})" for c in countries], 1)
    country = countries[ci]

    adv = ask_yes_no("2)  Are payments made in ADVANCE (start of period)?\n"
                     "    (No = arrears, i.e. end of period)", True)

    freqs = list(FREQUENCIES)
    fi = ask_choice("3)  How often are payments made?",
                    [f"{FREQUENCIES[f]}  ({f}/year)" for f in freqs], 0)
    freq = freqs[fi]

    c = COUNTRIES[country]
    print(f"""
  ── Configuration ─────────────────────────────
     Country      : {country}  ·  {c['cur']} ({c['sym']})
     Framework    : {c['framework']}  ·  {c['std']}
     Timing       : {"Advance" if adv else "Arrears"}
     Frequency    : {FREQUENCIES[freq]} ({freq}/year)
     Audit refs   : {c['isa540']} · {c['isa500']} · {c['isa230']}
  ──────────────────────────────────────────────""")
    return {"country": country, "is_advance": adv, "freq": freq}


# ═══════════════════════════════════════════════════════════════════════
#  SHARED — the LIQUID Inputs sheet
#  (used by BOTH the input template and the calculation workbook)
# ═══════════════════════════════════════════════════════════════════════

CURRENCY_ARRAY = ('{"Türkiye","TRY";"Ireland","EUR";"United Kingdom","GBP";'
                  '"UAE","AED";"Australia","AUD"}')


def _sect(ws, row, text):
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_TITLE
    c.fill = NAVY_FILL
    for col in (2, 3):
        ws.cell(row=row, column=col).fill = NAVY_FILL


def _inp(ws, row, value, numfmt="General"):
    """Dark-red editable input cell."""
    c = ws.cell(row=row, column=2, value=value)
    c.fill = RED_FILL
    c.font = F_INPUT
    c.border = BOX_M
    c.number_format = numfmt
    c.alignment = Alignment(horizontal="center")
    return c


def _auto(ws, row, formula, numfmt="General", link=False, check=False):
    """Black formula cell (green if link / check)."""
    c = ws.cell(row=row, column=2, value=formula)
    c.font = F_LINK if link else (F_CHECK if check else F_LABEL)
    if link or check:
        c.fill = GREEN_FILL
    c.number_format = numfmt
    c.alignment = Alignment(horizontal="center")
    return c


def _lbl(ws, row, label, note=""):
    ws.cell(row=row, column=1, value=label).font = F_LABEL
    if note:
        ws.cell(row=row, column=3, value=note).font = F_NOTE


def build_inputs_sheet(ws, inp: LeaseInputs, calculation_file: bool,
                       blank_required: bool = False) -> None:
    """Writes the Inputs sheet.  calculation_file=True adds the live links
    to the other sheets (section 5) and the Setup VLOOKUPs (section 6)."""
    ws.sheet_view.showGridLines = False

    ws["A1"] = "IFRS 16  —  Lessee Accounting for a Lease  (LIQUID model)"
    ws["A1"].font = F_TITLE
    ws["A1"].fill = NAVY_FILL
    ws["B1"].fill = NAVY_FILL
    ws["C1"].fill = NAVY_FILL
    ws["A2"] = ("Single-asset re-performance model. Edit only the DARK-RED "
                "input cells (thick border). Everything else auto-calculates.")
    ws["A2"].font = F_NOTE
    ws["A3"] = ("KEY:  dark red filling cells = your input   ·   black = formula"
                "   ·   green = link to another sheet")
    ws["A3"].font = F_NOTE

    # In blank mode the REQUIRED inputs are left empty so the tool cannot
    # silently run on example values — validation forces a real fill.
    req = (lambda v: None) if blank_required else (lambda v: v)

    # ── 1. Lease terms ──────────────────────────
    _sect(ws, 4, "1.  Lease terms")
    ws["C4"] = "Notes"
    ws["C4"].font = F_HDR
    ws["C4"].fill = NAVY_FILL

    _lbl(ws, 5, "Lessee entity");                    _inp(ws, 5, req(inp.entity))
    _lbl(ws, 6, "Underlying asset");                 _inp(ws, 6, req(inp.asset))
    _lbl(ws, 7, "Reporting currency", "auto from the Country selector (row 48)")
    if calculation_file:
        _auto(ws, 7, '=VLOOKUP($B$48,Setup!$A$2:$I$6,2,0)&" ("'
                     '&VLOOKUP($B$48,Setup!$A$2:$I$6,3,0)&")"', link=True)
    else:
        _auto(ws, 7, f'=IFERROR(VLOOKUP($B$48,{CURRENCY_ARRAY},2,0),"EUR")')
    _lbl(ws, 8, "Lease commencement date", "First day the asset is available for use")
    _inp(ws, 8, req(inp.commencement), FMT_DATE)
    _lbl(ws, 9, "Lease term — YEARS", "◀ enter whole years (e.g. 5)")
    _inp(ws, 9, req(inp.term_years), "0")
    _lbl(ws, 10, "Lease term — plus extra MONTHS", "◀ optional, e.g. 6 for 3½ years")
    _inp(ws, 10, inp.term_extra_months, "0")
    _lbl(ws, 11, "Total lease term (months)", "auto: years × 12 + months  ← liquid")
    _auto(ws, 11, "=B9*12+B10", "0")
    _lbl(ws, 12, "Payment frequency (periods / year)",
         "12=monthly · 4=quarterly · 2=semi-annual · 1=annual")
    _inp(ws, 12, inp.freq, "0")
    _lbl(ws, 13, "Months per period", "auto: 12 ÷ frequency")
    _auto(ws, 13, "=12/B12", "0")
    _lbl(ws, 14, "Number of payment periods (N)", "auto: drives the whole schedule")
    _auto(ws, 14, "=B11/B13", "0")
    _lbl(ws, 15, "Lease payment per period", "Fixed contractual payment each period")
    _inp(ws, 15, req(inp.payment), FMT_MONEY0)
    _lbl(ws, 16, "Payment timing",
         "Advance = paid at start of period (typical for vehicles)")
    _inp(ws, 16, "Advance" if inp.is_advance else "Arrears")
    _lbl(ws, 17, "Timing flag (1=advance, 0=arrears)", "auto from the dropdown above")
    _auto(ws, 17, '=IF(B16="Advance",1,0)', "0")

    # ── 2. Discount rate ────────────────────────
    _sect(ws, 19, "2.  Discount rate")
    _lbl(ws, 20, "Annual incremental borrowing rate (IBR)",
         "Rate to borrow similar funds, similar term/security")
    _inp(ws, 20, req(inp.ibr_annual), FMT_PCT2)
    _lbl(ws, 21, "Periodic discount rate (per period)",
         "auto: effective conversion of the annual rate")
    _auto(ws, 21, "=(1+B20)^(1/B12)-1", FMT_PCT4)

    # ── 3. Other initial-measurement inputs ─────
    _sect(ws, 23, "3.  Other initial-measurement inputs")
    _lbl(ws, 24, "Initial direct costs (IDC)", "e.g. delivery, registration, legal")
    _inp(ws, 24, inp.idc, FMT_MONEY0)
    _lbl(ws, 25, "Prepaid lease payments before commencement")
    _inp(ws, 25, inp.prepaid, FMT_MONEY0)
    _lbl(ws, 26, "Lease incentives received", "Reduces the ROU asset")
    _inp(ws, 26, inp.incentives, FMT_MONEY0)
    _lbl(ws, 27, "Estimated restoration / dismantling cost",
         "Enter the already-discounted provision, if any")
    _inp(ws, 27, inp.restoration, FMT_MONEY0)

    # ── 4. Depreciation ─────────────────────────
    _sect(ws, 29, "4.  Depreciation of the right-of-use asset")
    _lbl(ws, 30, "Method", "Choose Straight-line or Reducing balance")
    _inp(ws, 30, inp.method)
    _lbl(ws, 31, "Useful life of the asset (months)", "◀ key assumption")
    _inp(ws, 31, req(inp.useful_life_months), "0")
    _lbl(ws, 32, "Useful life (in periods)")
    _auto(ws, 32, "=B31/B13", "0")
    _lbl(ws, 33, "Depreciation period (in periods)",
         "shorter of lease term and useful life (no ownership transfer)")
    _auto(ws, 33, "=MIN(B14,INT(B32))", "0")
    _lbl(ws, 34, "Reducing-balance annual rate (if used)",
         "only used when Method = Reducing balance")
    _inp(ws, 34, inp.rb_rate, FMT_PCT2)

    # ── 5. Key outputs ──────────────────────────
    if calculation_file:
        _sect(ws, 35, "5.  Key outputs  (calculated)")
        outs = [
            ("Lease liability at commencement",   "='Initial Measurement'!E4", FMT_MONEY),
            ("Right-of-use asset at commencement", "='Initial Measurement'!N9", FMT_MONEY),
            ("Number of payment periods",          "=B14",                      "0"),
            ("Periodic discount rate",             "=B21",                      FMT_PCT4),
            ("Total interest over lease term",     "='Lease Schedule'!E2",      FMT_MONEY),
            ("Total depreciation over lease term", "='Lease Schedule'!I2",      FMT_MONEY),
            ("Total expense over lease term",      "='Lease Schedule'!K2",      FMT_MONEY),
        ]
        for i, (label, f, fmt) in enumerate(outs):
            _lbl(ws, 36 + i, label)
            _auto(ws, 36 + i, f, fmt, link=True)
    else:
        _sect(ws, 35, "5.  Key outputs")
        ws["A36"] = ("Outputs appear in the generated IFRS16_Calculation.xlsx "
                     "after you return to the script.")
        ws["A36"].font = F_NOTE

    _lbl(ws, 44, "Template capacity check")
    _auto(ws, 44, f'=IF(B14<={MAX_PERIODS},"OK","TERM TOO LONG: max '
                  f'{MAX_PERIODS} periods")', check=True)
    _lbl(ws, 45, "Whole-number-of-periods check")
    _auto(ws, 45, '=IF(B14=INT(B14),"OK","ADJUST term/frequency: '
                  'N is not a whole number")', check=True)

    # ── 6. Country & reporting framework ────────
    _sect(ws, 47, "6.  Country & reporting framework")
    _lbl(ws, 48, "Country / jurisdiction",
         "drives currency, framework, audit standards, tax rate")
    _inp(ws, 48, inp.country)
    frame_rows = [
        ("Reporting framework",       4, "General"),
        ("Lease standard",            5, "General"),
        ("Auditing — estimates",      6, "General"),
        ("Auditing — evidence",       7, "General"),
        ("Auditing — documentation",  8, "General"),
    ]
    for i, (label, col, fmt) in enumerate(frame_rows):
        _lbl(ws, 49 + i, label)
        if calculation_file:
            _auto(ws, 49 + i, f"=VLOOKUP($B$48,Setup!$A$2:$I$6,{col},0)",
                  fmt, link=True)
        else:
            ws.cell(row=49 + i, column=2, value="auto in calculation file").font = F_NOTE
    _lbl(ws, 54, "Corporate tax rate" + (" (illustrative)" if calculation_file else ""))
    if calculation_file:
        _auto(ws, 54, "=VLOOKUP($B$48,Setup!$A$2:$I$6,9,0)", FMT_PCT2, link=True)
    else:
        ws.cell(row=54, column=2, value="auto in calculation file").font = F_NOTE

    # ── dropdowns ───────────────────────────────
    for formula, cell in (('"12,4,2,1"', "B12"),
                          ('"Advance,Arrears"', "B16"),
                          ('"Straight-line,Reducing balance"', "B30"),
                          ('"Türkiye,Ireland,United Kingdom,UAE,Australia"', "B48")):
        dv = DataValidation(type="list", formula1=formula, allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(ws[cell])

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 46


# ═══════════════════════════════════════════════════════════════════════
#  STEP 2 — BUILD THE INPUT TEMPLATE  (single Inputs sheet)
# ═══════════════════════════════════════════════════════════════════════

def build_template(cfg: dict, path: str, prefill: bool = False) -> None:
    """Write the input template.  By default the required cells (entity,
    date, years, payment, IBR, useful life) are BLANK so the tool cannot
    run on example values.  prefill=True writes the golden-case example
    (used by demo())."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    preset = LeaseInputs(country=cfg["country"], is_advance=cfg["is_advance"],
                         freq=cfg["freq"])
    build_inputs_sheet(ws, preset, calculation_file=False,
                       blank_required=not prefill)
    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
#  STEP 3 — READ THE FILLED TEMPLATE BACK
# ═══════════════════════════════════════════════════════════════════════

def read_template(path: str) -> LeaseInputs:
    wb = load_workbook(path, data_only=True)
    if "Inputs" not in wb.sheetnames:
        raise ValueError("Sheet 'Inputs' not found — is this the right file?")
    ws = wb["Inputs"]
    g = lambda cell: ws[cell].value

    errors = []

    entity = str(g("B5") or "").strip()
    asset  = str(g("B6") or "").strip()
    if not entity:
        errors.append("Lessee entity (B5) is empty")

    raw_date = g("B8")
    if isinstance(raw_date, datetime):
        commencement = raw_date.date()
    elif isinstance(raw_date, date):
        commencement = raw_date
    else:
        try:
            commencement = datetime.strptime(
                str(raw_date).strip()[:10], "%Y-%m-%d").date()
        except (ValueError, AttributeError):
            errors.append("Commencement date (B8) is empty" if raw_date is None
                          else f"Commencement date (B8) must be a date — got '{raw_date}'")
            commencement = date.today()

    def num(cell, label, required=True):
        v = g(cell)
        if v is None or str(v).strip() == "":
            if required:
                errors.append(f"{label} ({cell}) is empty")
            return 0.0
        try:
            return float(v)
        except (TypeError, ValueError):
            errors.append(f"{label} ({cell}) is not a number — got '{v}'")
            return 0.0

    years       = int(num("B9",  "Lease term — YEARS"))
    extra       = int(num("B10", "Extra months", required=False))
    freq        = int(num("B12", "Payment frequency"))
    payment     = num("B15", "Lease payment")
    ibr         = num("B20", "Annual IBR")
    idc         = num("B24", "Initial direct costs", required=False)
    prepaid     = num("B25", "Prepaid payments", required=False)
    incentives  = num("B26", "Lease incentives", required=False)
    restoration = num("B27", "Restoration provision", required=False)
    useful_life = int(num("B31", "Useful life"))
    rb_rate     = num("B34", "Reducing-balance rate", required=False) or 0.30

    timing  = str(g("B16") or "Advance").strip().capitalize()
    method  = str(g("B30") or "Straight-line").strip()
    method  = {"straight-line": "Straight-line",
               "reducing balance": "Reducing balance"}.get(method.lower(), method)
    country = str(g("B48") or "Ireland").strip()
    country = {c.lower(): c for c in COUNTRIES}.get(country.lower(), country)

    if country not in COUNTRIES:
        errors.append(f"Country (B48) must be one of {list(COUNTRIES)} — got '{country}'")
    if freq not in FREQUENCIES:
        errors.append(f"Frequency (B12) must be 12/4/2/1 — got '{freq}'")
    if method not in ("Straight-line", "Reducing balance"):
        errors.append(f"Method (B30) must be Straight-line or Reducing balance — got '{method}'")

    term = years * 12 + extra
    if term <= 0:            errors.append("Lease term must be > 0 (B9/B10)")
    if payment <= 0:         errors.append("Lease payment (B15) must be > 0")
    if ibr >= 1:             errors.append("IBR (B20) looks like a percentage — enter 0.06 or 6%")
    elif not (0 < ibr < 1):  errors.append("IBR (B20) must be between 0 and 1")
    if useful_life <= 0:     errors.append("Useful life (B31) must be > 0")
    if timing not in ("Advance", "Arrears"):
        errors.append(f"Payment timing (B16) must be Advance or Arrears — got '{g('B16')}'")

    if freq in FREQUENCIES:
        mpp = 12 / freq
        n   = term / mpp
        if n != int(n):
            errors.append(f"Term ({term} months) does not give a whole number of "
                          f"{FREQUENCIES[freq].lower()} periods (N = {n})")
        elif n > MAX_PERIODS:
            errors.append(f"N = {int(n)} periods exceeds the template capacity "
                          f"of {MAX_PERIODS}")

    if errors:
        raise ValueError("Template validation failed:\n  • " + "\n  • ".join(errors))

    return LeaseInputs(
        entity=entity, asset=asset, country=country,
        commencement=commencement, term_years=years, term_extra_months=extra,
        freq=freq, payment=payment, is_advance=(timing == "Advance"),
        ibr_annual=ibr, idc=idc, prepaid=prepaid, incentives=incentives,
        restoration=restoration, method=method,
        useful_life_months=useful_life, rb_rate=rb_rate,
    )


# ═══════════════════════════════════════════════════════════════════════
#  STEP 4 — BUILD THE LIQUID CALCULATION WORKBOOK  (10 sheets, all formulas)
# ═══════════════════════════════════════════════════════════════════════

def _title(ws, text, ncols=1):
    ws["A1"] = text
    ws["A1"].font = F_TITLE
    for col in range(1, ncols + 1):
        ws.cell(row=1, column=col).fill = NAVY_FILL


def _hdr(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = F_HDR
        c.fill = NAVY_FILL
        c.alignment = Alignment(horizontal="center")


def _f(ws, row, col, formula, numfmt=FMT_MONEY, font=F_LABEL, fill=None):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = numfmt
    c.font = font
    if fill:
        c.fill = fill
    return c


def build_calculation_workbook(inp: LeaseInputs, path: str) -> None:
    wb = Workbook()

    # ═══════════ 1. INPUTS (with live links + Setup lookups) ═══════════
    ws = wb.active
    ws.title = "Inputs"
    build_inputs_sheet(ws, inp, calculation_file=True)

    # ═══════════ 2. INITIAL MEASUREMENT ═══════════
    ws = wb.create_sheet("Initial Measurement")
    ws.sheet_view.showGridLines = False
    _title(ws, "Initial Measurement at Commencement", ncols=14)

    ws["A3"] = "A.  Lease liability  =  Present value of the lease payments"
    ws["A3"].font = F_HDR
    for col in range(1, 6):
        ws.cell(row=3, column=col).fill = NAVY_FILL
    ws["K3"] = "B.  Right-of-use (ROU) asset at commencement"
    ws["K3"].font = F_HDR
    for col in range(11, 15):
        ws.cell(row=3, column=col).fill = NAVY_FILL

    # totals row 4
    ws["B4"] = "Total"
    ws["B4"].font = F_BOLD
    _f(ws, 4, 3, "=SUM(C6:C48576)", FMT_MONEY, F_BOLD)
    _f(ws, 4, 4, "=SUM(D6:D48576)", FMT_FACTOR, F_BOLD)
    _f(ws, 4, 5, "=SUM(E6:E48576)", FMT_MONEY, F_BOLD, GREEN_FILL)

    # ROU build-up (K4:N9)
    rou_rows = [
        ("Lease liability at commencement",     "=E4"),
        ("Add:  Initial direct costs",           "=Inputs!$B$24"),
        ("Add:  Prepaid lease payments",         "=Inputs!$B$25"),
        ("Less: Lease incentives received",      "=-Inputs!$B$26"),
        ("Add:  Estimated restoration provision", "=Inputs!$B$27"),
    ]
    for i, (label, f) in enumerate(rou_rows):
        ws.cell(row=4 + i, column=11, value=label).font = F_BODY
        _f(ws, 4 + i, 14, f)
    ws.cell(row=9, column=11, value="Right-of-use asset at commencement").font = F_BOLD
    _f(ws, 9, 14, "=SUM(N4:N8)", FMT_MONEY, F_BOLD, GREEN_FILL)

    _hdr(ws, 5, ["Period", "Payment date", "Lease payment",
                 "Discount factor", "Present value"])
    for r in range(6, 6 + MAX_PERIODS):
        _f(ws, r, 1, f'=IF((ROW()-5)<=Inputs!$B$14,ROW()-5,"")', "0")
        _f(ws, r, 2, f'=IF($A{r}="","",EDATE(Inputs!$B$8,'
                     f'($A{r}-Inputs!$B$17)*Inputs!$B$13))', FMT_DATE)
        _f(ws, r, 3, f'=IF($A{r}="","",Inputs!$B$15)')
        _f(ws, r, 4, f'=IF($A{r}="","",1/(1+Inputs!$B$21)^($A{r}-Inputs!$B$17))',
           FMT_FACTOR)
        _f(ws, r, 5, f'=IF($A{r}="","",C{r}*D{r})')

    for col, w in zip("ABCDE", [10, 17.8, 19, 19.3, 17.5]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["K"].width = 40
    ws.column_dimensions["L"].width = 33
    ws.column_dimensions["N"].width = 14

    # ═══════════ 3. LEASE SCHEDULE ═══════════
    ws = wb.create_sheet("Lease Schedule")
    ws.sheet_view.showGridLines = False
    _title(ws, "Lease Liability Amortisation  &  ROU Depreciation Schedule",
           ncols=13)

    ws["C2"] = "Total"
    ws["C2"].font = F_BOLD
    for col in range(4, 12):
        letter = get_column_letter(col)
        _f(ws, 2, col, f"=SUM({letter}4:{letter}1048576)", FMT_MONEY, F_BOLD)

    _hdr(ws, 3, ["Period", "Payment date", "Opening liability", "Lease payment",
                 "Interest expense", "Principal repaid", "Closing liability",
                 "Opening ROU", "Depreciation", "Closing ROU", "Total expense",
                 "Year", "Payment Status"])

    depr_formula = ('=IF($A{r}="","",IF($A{r}>Inputs!$B$33,0,'
                    'IF(Inputs!$B$30="Straight-line",'
                    "'Initial Measurement'!$N$9/Inputs!$B$33,"
                    'IF($A{r}=Inputs!$B$33,H{r},'
                    'H{r}*(1-(1-Inputs!$B$34)^(1/Inputs!$B$12))))))')

    for r in range(4, 4 + MAX_PERIODS):
        _f(ws, r, 1, f'=IF((ROW()-3)<=Inputs!$B$14,ROW()-3,"")', "0")
        _f(ws, r, 2, f'=IF($A{r}="","",EDATE(Inputs!$B$8,'
                     f'($A{r}-Inputs!$B$17)*Inputs!$B$13))', FMT_DATE)
        if r == 4:
            _f(ws, r, 3, "='Initial Measurement'!$E$4")
            _f(ws, r, 8, "='Initial Measurement'!$N$9")
        else:
            _f(ws, r, 3, f'=IF($A{r}="","",IF($A{r}=1,'
                         f"'Initial Measurement'!$E$4,G{r-1}))")
            _f(ws, r, 8, f'=IF($A{r}="","",J{r-1})')
        _f(ws, r, 4, f'=IF($A{r}="","",Inputs!$B$15)')
        _f(ws, r, 5, f'=IF($A{r}="","",(C{r}-Inputs!$B$17*D{r})*Inputs!$B$21)')
        _f(ws, r, 6, f'=IF($A{r}="","",D{r}-E{r})')
        _f(ws, r, 7, f'=IF($A{r}="","",C{r}-D{r}+E{r})')
        _f(ws, r, 9, depr_formula.format(r=r))
        _f(ws, r, 10, f'=IF($A{r}="","",H{r}-I{r})')
        _f(ws, r, 11, f'=IF($A{r}="","",E{r}+I{r})')
        _f(ws, r, 12, f'=IF($A{r}="","",YEAR(EDATE(Inputs!$B$8,'
                      f'($A{r}-1)*Inputs!$B$13)))', "0")
        _f(ws, r, 13, f'=IF(N(H{r})=0,"",IF(ROUND(G{r},2)=0,'
                      f'"Final Payment","Continuing"))', "General")

    for col, w in zip("ABCDEFGHIJKLM",
                      [11.5, 17.3, 19.7, 18.2, 20, 19.8, 18.8,
                       17.7, 16.5, 16.8, 17.2, 10.2, 17.7]):
        ws.column_dimensions[col].width = w

    # ═══════════ 4. ANNUAL SUMMARY ═══════════
    ws = wb.create_sheet("Annual Summary")
    ws.sheet_view.showGridLines = False
    _title(ws, "Annual Summary  —  P&L Impact and Year-End Balances", ncols=15)

    ws["A3"] = "Profit-or-loss impact and closing balances by reporting year"
    ws["A3"].font = F_BOLD
    ws["H3"] = "Memo - IFRS 16 expense profile vs. a straight-line charge"
    ws["H3"].font = F_BOLD
    ws["M3"] = "Reconciliation checks"
    ws["M3"].font = F_BOLD
    ws["N3"] = "Balance"
    ws["N3"].font = F_BOLD
    ws["O3"] = "Status"
    ws["O3"].font = F_BOLD

    ws["A4"] = "Total"
    ws["A4"].font = F_BOLD
    for col in "BCDEF":
        _f(ws, 4, ws[col + "4"].column, f"=SUM({col}6:{col}125)",
           FMT_MONEY, F_BOLD)
    ws["H4"] = "Total"
    ws["H4"].font = F_BOLD
    for col in "IJK":
        _f(ws, 4, ws[col + "4"].column, f"=SUM({col}6:{col}125)",
           FMT_MONEY, F_BOLD)

    _hdr(ws, 5, ["Reporting year", "Interest expense", "Depreciation",
                 "Total IFRS 16 expense", "Lease liability (year-end)",
                 "ROU asset (year-end)"])
    _hdr(ws, 5, ["Year", "IFRS 16 total expense"], start_col=8)
    c = ws.cell(row=5, column=10, value='=Inputs!$B$30&" equivalent"')
    c.font = F_HDR
    c.fill = NAVY_FILL
    _hdr(ws, 5, ["Front-loading (IFRS 16 − SL)"], start_col=11)

    for r in range(6, 126):
        if r == 6:
            _f(ws, r, 1, "=MIN('Lease Schedule'!$L$4:$L$123)", "0")
        else:
            _f(ws, r, 1, f'=IF(A{r-1}="","",IF(COUNTIF('
                         f"'Lease Schedule'!$L$4:$L$123,A{r-1}+1)>0,"
                         f'A{r-1}+1,""))', "0")
        _f(ws, r, 2, f'=IF($A{r}="","",SUMIF(\'Lease Schedule\'!$L$4:$L$123,'
                     f"$A{r},'Lease Schedule'!$E$4:$E$123))")
        _f(ws, r, 3, f'=IF($A{r}="","",SUMIF(\'Lease Schedule\'!$L$4:$L$123,'
                     f"$A{r},'Lease Schedule'!$I$4:$I$123))")
        _f(ws, r, 4, f'=IF($A{r}="","",B{r}+C{r})')
        _f(ws, r, 5, f'=IF($A{r}="","",LOOKUP(2,1/(\'Lease Schedule\'!'
                     f"$L$4:$L$123=$A{r}),'Lease Schedule'!$G$4:$G$123))")
        _f(ws, r, 6, f'=IF($A{r}="","",LOOKUP(2,1/(\'Lease Schedule\'!'
                     f"$L$4:$L$123=$A{r}),'Lease Schedule'!$J$4:$J$123))")
        _f(ws, r, 8, f"=A{r}", "0")
        _f(ws, r, 9, f'=IF($A{r}="","",D{r})')
        _f(ws, r, 10, f'=IF(H{r}="","",SUM($I$6:$I$125)/COUNT($H$6:$H$125))')
        _f(ws, r, 11, f'=IF($A{r}="","",I{r}-J{r})')

    checks = [
        ("Closing lease liability at end of term",
         "=LOOKUP(2,1/('Lease Schedule'!$A$4:$A$123<>\"\"),"
         "'Lease Schedule'!$G$4:$G$123)"),
        ("Closing ROU asset at end of term",
         "=LOOKUP(2,1/('Lease Schedule'!$A$4:$A$123<>\"\"),"
         "'Lease Schedule'!$J$4:$J$123)"),
        ("Total principal repaid − initial liability",
         "='Lease Schedule'!F2-'Initial Measurement'!E4"),
        ("Total expense − (total payments + IDC)",
         "='Lease Schedule'!K2-'Lease Schedule'!D2-Inputs!$B$24"),
        ("Sum of front-loading differences", "=K4"),
    ]
    for i, (label, f) in enumerate(checks):
        r = 6 + i
        ws.cell(row=r, column=13, value=label).font = F_BODY
        _f(ws, r, 14, f)
        _f(ws, r, 15, f'=IF(ABS(N{r})<0.01,"OK","CHECK")', "General",
           F_CHECK, GREEN_FILL)

    for col, w in zip("ABCDEFGHIJKLMNO",
                      [15, 16, 14, 18, 20, 17, 3, 8, 17, 18, 20, 3, 36, 13, 8]):
        ws.column_dimensions[col].width = w

    # ═══════════ 5. JOURNALS ═══════════
    ws = wb.create_sheet("Journals")
    ws.sheet_view.showGridLines = False
    _title(ws, "Journal Entries", ncols=4)

    ws["A2"] = "1.  Initial recognition at commencement"
    ws["A2"].font = F_BOLD
    ws["A3"] = "Date of entry"
    ws["A3"].font = F_LABEL
    _f(ws, 3, 2, "=Inputs!$B$8", FMT_DATE)
    _hdr(ws, 4, ["Account", "Debit", "Credit", "Note"])

    je1 = [
        ("Right-of-use asset",
         "='Initial Measurement'!$N$9", None, "Dr — asset recognised"),
        ("Cash / bank — lease incentive received",
         '=IF(Inputs!$B$26>0,Inputs!$B$26,"-")', None,
         "Dr — incentive received, if any"),
        ("    Lease liability",
         None, "='Initial Measurement'!$E$4", "Cr — PV of lease payments"),
        ("    Cash / bank — initial direct costs",
         None, '=IF(Inputs!$B$24>0,Inputs!$B$24,"-")', "Cr — IDC paid"),
        ("    Cash / bank — prepaid lease payments",
         None, '=IF(Inputs!$B$25>0,Inputs!$B$25,"-")', "Cr — prepayments, if any"),
        ("    Restoration / dismantling provision",
         None, '=IF(Inputs!$B$27>0,Inputs!$B$27,"-")', "Cr — provision, if any"),
    ]
    for i, (acct, dr, cr, note) in enumerate(je1):
        r = 5 + i
        ws.cell(row=r, column=1, value=acct).font = F_BODY
        if dr: _f(ws, r, 2, dr)
        if cr: _f(ws, r, 3, cr)
        ws.cell(row=r, column=4, value=note).font = F_NOTE
    ws["A11"] = "Total"
    ws["A11"].font = F_BOLD
    _f(ws, 11, 2, '=SUMIF(B5:B10,">0")', FMT_MONEY, F_BOLD)
    _f(ws, 11, 3, '=SUMIF(C5:C10,">0")', FMT_MONEY, F_BOLD)
    ws["A12"] = "Balance check"
    ws["A12"].font = F_LABEL
    _f(ws, 12, 2, '=SUMIF(B5:B10,">0")-SUMIF(C5:C10,">0")')
    _f(ws, 12, 3, '=IF(ABS(B11-C11)<0.01,"OK, debits = credits","CHECK")',
       "General", F_CHECK, GREEN_FILL)

    ws["A14"] = "2.  Representative entries — Period 1"
    ws["A14"].font = F_BOLD
    ws["A15"] = ("Subsequent periods repeat this pattern using the "
                 "Lease Schedule amounts.")
    ws["A15"].font = F_NOTE
    _hdr(ws, 16, ["Account", "Debit", "Credit", "Note"])
    je2 = [
        ("(a)  Lease liability", "='Lease Schedule'!$D$4", None, "Dr — payment"),
        ("        Cash / bank", None, "='Lease Schedule'!$D$4", "Cr — cash paid"),
        ("(b)  Interest expense (finance cost)",
         "='Lease Schedule'!$E$4", None, "Dr — unwinding of discount"),
        ("        Lease liability", None, "='Lease Schedule'!$E$4",
         "Cr — increases the liability"),
        ("(c)  Depreciation expense", "='Lease Schedule'!$I$4", None,
         "Dr — ROU depreciation"),
        ("        Accumulated depreciation (ROU)", None,
         "='Lease Schedule'!$I$4", "Cr — reduces carrying amount"),
    ]
    for i, (acct, dr, cr, note) in enumerate(je2):
        r = 17 + i
        ws.cell(row=r, column=1, value=acct).font = F_BODY
        if dr: _f(ws, r, 2, dr)
        if cr: _f(ws, r, 3, cr)
        ws.cell(row=r, column=4, value=note).font = F_NOTE
    ws["A23"] = "Period 1 charge to profit or loss"
    ws["A23"].font = F_BOLD
    _f(ws, 23, 2, "='Lease Schedule'!$E$4+'Lease Schedule'!$I$4",
       FMT_MONEY, F_BOLD, GREEN_FILL)
    ws["C23"] = "Finance cost + depreciation"
    ws["C23"].font = F_NOTE

    for col, w in zip("ABCD", [42, 16, 16, 34]):
        ws.column_dimensions[col].width = w

    # ═══════════ 6. TAX_RECONCILIATION ═══════════
    ws = wb.create_sheet("Tax_Reconciliation")
    ws.sheet_view.showGridLines = False
    _title(ws, "Tax treatment — deferred-tax reconciliation", ncols=5)
    ws["A2"] = ("Tax rate is taken from Setup (changes with Country). "
                "Illustrative year-1 figures link to the Annual Summary.")
    ws["A2"].font = F_NOTE
    _f(ws, 3, 1, '="Jurisdiction: "&Inputs!$B$48&"  |  Tax rate: "'
                 '&TEXT(VLOOKUP(Inputs!$B$48,Setup!$A$2:$I$6,9,0),"0.0%")'
                 '&" illustrative"', "General", F_BOLD)
    ws["A4"] = ("Many jurisdictions that adopt IFRS align the lessee's tax "
                "deduction with the accounting (depreciation + interest); "
                "others deduct the cash rental — creating temporary differences.")
    ws["A4"].font = F_NOTE

    ws["A8"] = "P&L impact — accounting vs cash rental (illustrative, year 1)"
    ws["A8"].font = F_HDR
    for col in range(1, 6):
        ws.cell(row=8, column=col).fill = NAVY_FILL
    _hdr(ws, 9, ["Item", "Accounting", "Cash rental", "Difference", "Note"])
    tax_pl = [
        ("Interest on lease liability", "='Annual Summary'!B6", "=0",
         "P&L charge under IFRS 16"),
        ("Depreciation of ROU asset", "='Annual Summary'!C6", "=0",
         "P&L charge under IFRS 16"),
        ("Cash rental paid",
         "=SUMIF('Lease Schedule'!$L:$L,'Annual Summary'!$A$6,"
         "'Lease Schedule'!$D:$D)", "=B12", "Deductible under old rules"),
    ]
    for i, (label, b, c_, note) in enumerate(tax_pl):
        r = 10 + i
        ws.cell(row=r, column=1, value=label).font = F_BODY
        _f(ws, r, 2, b)
        _f(ws, r, 3, c_)
        _f(ws, r, 4, f"=B{r}-C{r}")
        ws.cell(row=r, column=5, value=note).font = F_NOTE
    ws["A13"] = "Net P&L difference (year 1)"
    ws["A13"].font = F_BOLD
    _f(ws, 13, 2, "=B10+B11-B12", FMT_MONEY, F_BOLD)
    _f(ws, 13, 4, "=D10+D11-D12", FMT_MONEY, F_BOLD)

    ws["A15"] = "Deferred tax — book vs tax base (year-1 closing)"
    ws["A15"].font = F_HDR
    for col in range(1, 6):
        ws.cell(row=15, column=col).fill = NAVY_FILL
    _hdr(ws, 16, ["Item", "Book base", "Tax base", "Temp difference", "DT @ rate"])
    ws["A17"] = "ROU asset"
    ws["A17"].font = F_BODY
    _f(ws, 17, 2, "='Annual Summary'!F6")
    _f(ws, 17, 3, "0" if False else 0, FMT_MONEY)
    _f(ws, 17, 4, "=B17-C17")
    _f(ws, 17, 5, "=D17*VLOOKUP(Inputs!$B$48,Setup!$A$2:$I$6,9,0)")
    ws["A18"] = "Lease liability"
    ws["A18"].font = F_BODY
    _f(ws, 18, 2, "=-'Annual Summary'!E6")
    _f(ws, 18, 3, 0, FMT_MONEY)
    _f(ws, 18, 4, "=B18-C18")
    _f(ws, 18, 5, "=D18*VLOOKUP(Inputs!$B$48,Setup!$A$2:$I$6,9,0)")
    ws["A19"] = "Net deferred tax (liability) / asset"
    ws["A19"].font = F_BOLD
    _f(ws, 19, 5, "=E17+E18", FMT_MONEY, F_BOLD, GREEN_FILL)
    ws["A21"] = ("(Tax base illustratively set to nil — the cash-deduction "
                 "regime. Under an IFRS-aligned regime, set tax base = book "
                 "base and the temporary difference is nil.)")
    ws["A21"].font = F_NOTE

    for col, w in zip("ABCDE", [40, 16, 14, 17, 14]):
        ws.column_dimensions[col].width = w

    # ═══════════ 7. FINDINGS ═══════════
    ws = wb.create_sheet("Findings")
    ws.sheet_view.showGridLines = False
    _title(ws, "Summary of findings & recommendations", ncols=6)
    _f(ws, 2, 4, '="Amounts in "&Inputs!$B$7', "General", F_NOTE)
    _hdr(ws, 3, ["#", "Finding", "Description", "Amount",
                 "Adjustment", "Recommendation"])
    findings = [
        ("Example — discount rate stale",
         "IBR last reviewed 2022; current market rates higher. "
         "Risk ROU/liability overstated.",
         "=IF('Annual Summary'!$O$6=\"OK\",0,'Annual Summary'!$N$6)",
         "Investigate and adjust the lease calculation."),
        ("Example — embedded lease",
         "IT outsourcing for dedicated rack space meets IFRS 16.",
         "=IF('Annual Summary'!$O$7=\"OK\",0,'Annual Summary'!$N$7)",
         "Review IFRS 16 recognition and restate if material."),
        ("Example — disclosure gap",
         "Maturity analysis discloses undiscounted cash flows but "
         "reconciliation to liability is missing.",
         "=IF(Tax_Reconciliation!E19=0,0,Tax_Reconciliation!E19)",
         "Update disclosures and reconcile supporting schedules."),
    ]
    for i, (finding, desc, amt, rec) in enumerate(findings):
        r = 4 + i
        ws.cell(row=r, column=1, value=i + 1).font = F_BODY
        ws.cell(row=r, column=2, value=finding).font = F_BODY
        ws.cell(row=r, column=3, value=desc).font = F_BODY
        _f(ws, r, 4, amt)
        _f(ws, r, 5, f'=IF(ABS(D{r})>1,"Y","N")', "General")
        _f(ws, r, 6, f'=IF(E{r}="N","No adjustment required.","{rec}")', "General")
    for col, w in zip("ABCDEF", [5, 28, 52, 14, 12, 42]):
        ws.column_dimensions[col].width = w

    # ═══════════ 8. AUDIT_PROCEDURES ═══════════
    ws = wb.create_sheet("Audit_Procedures")
    ws.sheet_view.showGridLines = False
    _hdr(ws, 1, ["Ref", "Type", "Procedure", "Assertion", "Done by", "Date"])
    procedures = [
        ("P-1", "Existence", "Select sample of leases; vouch to signed contract "
         "& physical inspection (site/photo for property, VIN check for vehicles).", "EX"),
        ("P-2", "Completeness", "Obtain AP listing; filter recurring rent/lease "
         "vendors > threshold; reconcile each to the lease register.", "CO"),
        ("P-3", "Completeness", "Read board minutes and CFO Q&A for new property "
         "decisions, fleet expansion, equipment additions.", "CO"),
        ("P-4", "Completeness", "Obtain procurement contracts > threshold; assess "
         "whether any meet IFRS 16 §9 (identified asset + right to control use).", "CO"),
        ("P-5", "Accuracy", "Recompute PV of lease payments for sample using "
         "independently derived IBR; investigate diffs > threshold.", "AV"),
        ("P-6", "Accuracy — IBR", "Build benchmark IBR: risk-free curve + entity "
         "credit spread, adjusted for collateral & term. Compare to client rate.", "AV"),
        ("P-7", "Accuracy", "Recompute depreciation (ROU / shorter of useful life "
         "or term, straight-line). Agree to GL & FA register.", "AV"),
        ("P-8", "Cut-off", "Sample payments either side of year-end; ensure "
         "principal/interest split posted in correct period.", "CO"),
        ("P-9", "Classification", "Test current vs non-current split (next-12m "
         "principal). Verify ROU presented/disclosed per IFRS 16 §47.", "CL"),
        ("P-10", "Presentation", "Tick disclosure draft to checklist. Verify "
         "maturity analysis foots to total liability.", "PD"),
        ("P-11", "Lease term", "Challenge term assessment; review break/renewal "
         "options; corroborate with business plans & past practice.", "AV"),
        ("P-12", "Modifications", "Inspect modifications log; recompute "
         "remeasurement using revised discount rate where not a separate lease.", "AV"),
        ("P-13", "Variable payments", "Test CPI-linked rents: confirm remeasurement "
         "only at next reset date using the rate at that date (§42).", "AV"),
        ("P-14", "Expedients", "Test short-term & low-value expense; verify term "
         "≤12m (no purchase option); low-value below threshold.", "AV"),
        ("P-15", "FX leases", "Retranslate non-functional-currency liability at "
         "closing rate (monetary, IAS 21); ROU at historical (non-monetary).", "AV"),
        ("P-16", "Impairment", "Apply IAS 36 — consider impairment indicators on "
         "ROU (vacant property, abandoned site, restructuring).", "AV"),
        ("P-17", "Subleases", "Classify finance vs operating from head-lease ROU "
         "perspective (§B58). Recompute net investment where finance.", "CL"),
        ("P-18", "Sale-and-leaseback", "Apply IFRS 15 control test first; then "
         "IFRS 16 §99-103 measurement.", "AV"),
        ("P-19", "Tax", "Verify local tax treatment; reconcile any deferred tax "
         "(see Tax_Reconciliation).", "AV"),
        ("P-20", "Going concern", "Consider lease commitments' impact on liquidity "
         "& covenants — input to going-concern assessment.", "PD"),
    ]
    for i, (ref, typ, proc, assertion) in enumerate(procedures):
        r = 2 + i
        for col, val in enumerate((ref, typ, proc, assertion), start=1):
            ws.cell(row=r, column=col, value=val).font = F_BODY
    for col, w in zip("ABCDEF", [7, 18, 95, 10, 10, 10]):
        ws.column_dimensions[col].width = w

    # ═══════════ 9. PBC_LIST ═══════════
    ws = wb.create_sheet("PBC_List")
    ws.sheet_view.showGridLines = False
    _title(ws, "Prepared-By-Client (PBC) list for IFRS 16", ncols=6)
    _hdr(ws, 2, ["#", "Item requested", "Purpose / what the auditor does with it",
                 "Format", "Due", "Received"])
    pbc = [
        ("Complete lease register (Excel) with key terms per lease",
         "Foots to liability & ROU; basis for sampling", "XLSX", "T-15"),
        ("Signed lease contracts (PDF) for sampled leases",
         "Vouch terms: start, rent, escalations, options, breaks", "PDF", "T-10"),
        ("IBR derivation memo per currency / term band",
         "Recalculate discount rate; benchmark to risk-free + spread", "DOCX", "T-15"),
        ("Lease liability amortisation schedule (per lease & aggregate)",
         "Recompute interest & principal; agree to GL", "XLSX", "T-10"),
        ("ROU depreciation schedule (per lease & aggregate)",
         "Recompute depreciation; agree to FA register & GL", "XLSX", "T-10"),
        ("Movement schedule: opening → closing for ROU & liability",
         "Tie to PY balances; identify additions/mods/disposals", "XLSX", "T-10"),
        ("Modifications & reassessments log",
         "Verify triggers captured; recompute remeasurement", "XLSX", "T-10"),
        ("Short-term & low-value lease expense listing",
         "Test expedient eligibility; confirm P&L treatment", "XLSX", "T-10"),
        ("Variable lease payments listing (CPI, turnover rent)",
         "Verify exclusion from liability where appropriate", "XLSX", "T-10"),
        ("Procurement / AP listing for the year (full)",
         "Search for unrecorded leases — completeness", "CSV", "T-15"),
        ("Property / vehicle / IT asset registers",
         "Cross-check to lease register — completeness", "XLSX", "T-15"),
        ("Board / sub-committee minutes",
         "Identify new property decisions, exits, sale-and-leaseback", "PDF", "T-5"),
        ("Sub-lease contracts (if any)",
         "Lessor accounting — finance vs operating; net investment", "PDF", "T-10"),
        ("Sale-and-leaseback transactions log",
         "Test against IFRS 15 control & IFRS 16 §99-102", "DOCX", "T-10"),
        ("Draft IFRS 16 disclosures (note to FS)",
         "Tick to disclosure checklist", "DOCX", "T-5"),
        ("Management accounting-policy memo",
         "Confirm consistent application; identify policy choices", "DOCX", "T-15"),
        ("Deferred tax workings",
         "Confirm DTA/DTL on IFRS-vs-tax temporary difference", "XLSX", "T-5"),
        ("Foreign-currency leases — spot/closing rates",
         "Test FX retranslation of liability (monetary)", "XLSX", "T-5"),
    ]
    for i, (item, purpose, fmt, due) in enumerate(pbc):
        r = 3 + i
        ws.cell(row=r, column=1, value=i + 1).font = F_BODY
        ws.cell(row=r, column=2, value=item).font = F_BODY
        ws.cell(row=r, column=3, value=purpose).font = F_BODY
        ws.cell(row=r, column=4, value=fmt).font = F_BODY
        ws.cell(row=r, column=5, value=due).font = F_BODY
    for col, w in zip("ABCDEF", [5, 55, 55, 9, 8, 10]):
        ws.column_dimensions[col].width = w

    # ═══════════ 10. SETUP (country lookup) ═══════════
    ws = wb.create_sheet("Setup")
    ws.sheet_view.showGridLines = False
    _hdr(ws, 1, ["Country", "Currency", "Symbol", "Framework", "Lease standard",
                 "Audit — estimates", "Audit — evidence",
                 "Audit — documentation", "Tax rate"])
    for i, (country, c) in enumerate(COUNTRIES.items()):
        r = 2 + i
        vals = [country, c["cur"], c["sym"], c["framework"], c["std"],
                c["isa540"], c["isa500"], c["isa230"], c["tax"]]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = F_BODY
            if col == 9:
                cell.number_format = FMT_PCT2
    for col, w in zip("ABCDEFGHI", [16, 10, 8, 24, 14, 20, 20, 22, 10]):
        ws.column_dimensions[col].width = w

    wb.save(path)


# ═══════════════════════════════════════════════════════════════════════
#  CONSOLE SUMMARY (independent Python cross-check)
# ═══════════════════════════════════════════════════════════════════════

def print_summary(inp: LeaseInputs) -> None:
    chk = cross_check(inp)
    cur = COUNTRIES.get(inp.country, {}).get("cur", "")
    print(f"""
═══════════════════════════════════════════════════════════════════
  PYTHON CROSS-CHECK — {inp.entity}
  (independent re-computation of the Excel model, ISA 500 A21)
═══════════════════════════════════════════════════════════════════
  Periodic rate            : {chk['periodic_rate']:.10f}
  Lease Liability (Day 1)  : {cur} {chk['liability']:>14,.2f}   (IFRS 16, para 26)
  ROU Asset (Day 1)        : {cur} {chk['rou']:>14,.2f}   (IFRS 16, para 24)
  ─────────────────────────────────────────────────────────────────
  Total Interest Expense   : {cur} {chk['total_interest']:>14,.2f}
  Total Depreciation       : {cur} {chk['total_depreciation']:>14,.2f}
  Total P&L Charge         : {cur} {chk['total_interest'] + chk['total_depreciation']:>14,.2f}
  Total Payments           : {cur} {chk['total_payments']:>14,.2f}
  ─────────────────────────────────────────────────────────────────
  Closing liability        : {chk['closing_liability']:.6f}   {'✓ zero' if abs(chk['closing_liability']) < 1e-6 else '⚠ NON-ZERO'}
  Reconciliation: Liability + Interest = {chk['liability'] + chk['total_interest']:,.2f}
                  Total Payments       = {chk['total_payments']:,.2f}
                  {'✓ RECONCILED' if abs(chk['liability'] + chk['total_interest'] - chk['total_payments']) < 0.05 else '⚠ CHECK ROUNDING'}
═══════════════════════════════════════════════════════════════════
  Compare these figures against Inputs!B36:B42 of the generated
  workbook after Excel recalculates — they must agree to the cent.
═══════════════════════════════════════════════════════════════════""")


# ═══════════════════════════════════════════════════════════════════════
#  HIGH-LEVEL HELPERS — one-liners for Spyder / notebook use
# ═══════════════════════════════════════════════════════════════════════

def create_input(path: str = TEMPLATE_NAME, country: str = "Ireland",
                 advance: bool = True, freq: int = 12) -> str:
    """Write the Excel INPUT template and return its full path.

    Example (Spyder):
        import ifrs16kit as lease
        lease.create_input()                      # IFRS16_Input.xlsx
        # → open in Excel, fill the dark-red cells, save
    """
    if country not in COUNTRIES:
        raise ValueError(f"country must be one of {list(COUNTRIES)}")
    if freq not in FREQUENCIES:
        raise ValueError("freq must be 12, 4, 2 or 1")
    build_template({"country": country, "is_advance": advance, "freq": freq},
                   path)
    full = os.path.abspath(path)
    print(f"✓  Input template saved → {full}")
    print("   Open it in Excel, fill every DARK-RED cell, save, then run:")
    print(f"   lease.create_calculation('{path}')")
    return full


def create_calculation(template: str = TEMPLATE_NAME,
                       output: str = OUTPUT_NAME) -> str:
    """Read a filled input template, validate it, and write the Excel
    CALCULATION workbook (10 sheets, all live formulas).  Returns the path.

    Example (Spyder):
        lease.create_calculation()                # reads IFRS16_Input.xlsx
    """
    print(f"Reading inputs from → {os.path.abspath(template)}")
    inp = read_template(template)
    co = COUNTRIES[inp.country]
    print(f"""  ── Inputs read from the template ─────────────
     Entity        : {inp.entity}
     Country       : {inp.country}  ({co['cur']})
     Commencement  : {inp.commencement}
     Term          : {inp.term_years} years + {inp.term_extra_months} months = {inp.term_months} months → {int(inp.num_periods)} periods
     Frequency     : {FREQUENCIES[inp.freq]} · {"Advance" if inp.is_advance else "Arrears"}
     Payment       : {co['cur']} {inp.payment:,.2f}   ·   IBR {inp.ibr_annual:.2%}
     Depreciation  : {inp.method} · useful life {inp.useful_life_months} months
  ──────────────────────────────────────────────
  If any value above is NOT what you typed, you edited a different
  file than the one shown on the first line — check the path.""")
    print_summary(inp)
    build_calculation_workbook(inp, output)
    full = os.path.abspath(output)
    print(f"✓  Calculation workbook saved → {full}")
    print("   Open in Excel — it recalculates automatically.")
    return full


def demo(folder: str = ".") -> tuple:
    """Instantly write BOTH Excel files for the CALISMA golden case.
    Perfect first test that everything works.

    Example (Spyder):
        import ifrs16kit as lease
        lease.demo()
    """
    inp = LeaseInputs()                          # golden-case defaults
    # DEMO-suffixed names so the demo can NEVER overwrite your real
    # IFRS16_Input.xlsx / IFRS16_Calculation.xlsx
    t = os.path.join(folder, "IFRS16_Input_DEMO.xlsx")
    o = os.path.join(folder, "IFRS16_Calculation_DEMO.xlsx")
    build_template({"country": inp.country, "is_advance": inp.is_advance,
                    "freq": inp.freq}, t, prefill=True)
    print_summary(inp)
    build_calculation_workbook(inp, o)
    print(f"\n✓  Input template       → {os.path.abspath(t)}")
    print(f"✓  Calculation workbook → {os.path.abspath(o)}")
    print("   Expected in Excel: liability 14,761.09 · ROU 15,361.09")
    return os.path.abspath(t), os.path.abspath(o)


# ═══════════════════════════════════════════════════════════════════════
#  MAIN FLOW
# ═══════════════════════════════════════════════════════════════════════

def _confirm_and_run(inp: LeaseInputs) -> None:
    c = COUNTRIES[inp.country]
    print(f"""
  ── Extracted inputs ──────────────────────────
     Entity        : {inp.entity}
     Asset         : {inp.asset}
     Country       : {inp.country}  ({c['cur']})
     Commencement  : {inp.commencement}
     Term          : {inp.term_months} months → {int(inp.num_periods)} periods
     Frequency     : {FREQUENCIES[inp.freq]} · {"Advance" if inp.is_advance else "Arrears"}
     Payment       : {c['cur']} {inp.payment:,.2f}
     Annual IBR    : {inp.ibr_annual:.4%}
     IDC           : {c['cur']} {inp.idc:,.2f}
     Depreciation  : {inp.method} · useful life {inp.useful_life_months} months
  ──────────────────────────────────────────────""")
    if not ask_yes_no("  Run the re-performance with these inputs?", True):
        print("  Aborted — edit the template and run the script again.")
        return

    print_summary(inp)
    build_calculation_workbook(inp, OUTPUT_NAME)
    print(f"""
  ✓  LIQUID calculation workbook saved → {os.path.abspath(OUTPUT_NAME)}

     Sheets: Inputs · Initial Measurement · Lease Schedule
             Annual Summary · Journals · Tax_Reconciliation
             Findings · Audit_Procedures · PBC_List · Setup

     Every derived figure is a live Excel formula — open the file and
     Excel recalculates instantly.  Change any dark-red input cell
     (term, rate, country, method) and the whole model updates.

     Verify: Inputs!B36:B42 must equal the Python cross-check above.
""")


def main() -> None:
    if "--demo" in sys.argv:
        print("═" * 70)
        print("  DEMO MODE — CALISMA.xlsx golden case")
        print("═" * 70)
        # Route through demo(): it writes DEMO-suffixed files only, so the
        # CLI demo can never overwrite a real IFRS16_Calculation.xlsx.
        demo()
        print("     Expected: liability 14,761.09 · ROU 15,361.09 · interest 838.91")
        return

    # ── If a template already exists, offer to reuse it (don't wipe it!) ──
    if os.path.exists(TEMPLATE_NAME):
        print(f"\n  Found existing template: {os.path.abspath(TEMPLATE_NAME)}")
        if ask_yes_no("  Use this file (skip the 3 questions and read the "
                      "values already saved in it)?\n"
                      "  Answer 'n' to start fresh with a new blank template.", True):
            while True:
                try:
                    inp = read_template(TEMPLATE_NAME)
                    break
                except ValueError as e:
                    print(f"\n  ⚠ {e}\n")
                    input("  Fix the template, SAVE it, then press Enter... ")
                except PermissionError:
                    input("\n  ⚠ File is open in Excel — close it, then press Enter... ")
            _confirm_and_run(inp)
            return
        print("  OK — the existing template will be OVERWRITTEN.\n")

    cfg = interview()
    build_template(cfg, TEMPLATE_NAME)
    print(f"""
  ✓  Template created → {os.path.abspath(TEMPLATE_NAME)}

  ┌──────────────────────────────────────────────────────────┐
  │  NOW:                                                    │
  │  1. Open the file in Excel                               │
  │  2. Fill in every DARK-RED cell (entity, dates, term,    │
  │     payment, IBR, adjustments, method, useful life)      │
  │  3. Check both green validation cells say "OK"           │
  │  4. Save and close the file                              │
  │  5. Come back here and press Enter                       │
  └──────────────────────────────────────────────────────────┘
""")
    while True:
        input("  Press Enter when you have filled and SAVED the template... ")
        try:
            inp = read_template(TEMPLATE_NAME)
            break
        except ValueError as e:
            print(f"\n  ⚠ {e}\n")
            print("  Fix the template, save it, and press Enter again.\n")
        except PermissionError:
            print("\n  ⚠ The file is still open in Excel — close it first.\n")

    _confirm_and_run(inp)



# Interactive flow, importable:  lease.run()
#   1. asks the 3 questions in the console
#   2. writes IFRS16_Input.xlsx and PAUSES while you fill it in Excel
#   3. press Enter → validates → writes IFRS16_Calculation.xlsx
run = main


if __name__ == "__main__":
    main()

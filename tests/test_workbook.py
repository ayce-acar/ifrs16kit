"""Layer 3 — the generated 10-sheet calculation workbook.

Asserts the structural claims from the README: exactly ten sheets, every
derived figure a live formula ("liquid"), 120-period row-guarded schedule,
and a Setup sheet that mirrors the COUNTRIES config exactly.
"""
import pytest
from openpyxl import load_workbook

import ifrs16kit as kit

SHEETS = ["Inputs", "Initial Measurement", "Lease Schedule", "Annual Summary",
          "Journals", "Tax_Reconciliation", "Findings", "Audit_Procedures",
          "PBC_List", "Setup"]

# Sheets whose every cell must be a label or a formula — never a number.
FULLY_LIQUID = ["Initial Measurement", "Lease Schedule", "Annual Summary",
                "Journals"]


@pytest.fixture(scope="module")
def calc(tmp_path_factory):
    path = str(tmp_path_factory.mktemp("wb") / "IFRS16_Calculation.xlsx")
    kit.build_calculation_workbook(kit.LeaseInputs(), path)
    return load_workbook(path)


def test_exactly_ten_sheets_in_documented_order(calc):
    assert calc.sheetnames == SHEETS


def test_core_sheets_contain_no_hardcoded_numbers(calc):
    for name in FULLY_LIQUID:
        consts = [c.coordinate for row in calc[name].iter_rows() for c in row
                  if isinstance(c.value, (int, float))
                  and not isinstance(c.value, bool)]
        assert consts == [], f"{name} has baked-in numbers at {consts}"


def test_key_outputs_are_cross_sheet_links(calc):
    ws = calc["Inputs"]
    expected = {36: "'Initial Measurement'!E4", 37: "'Initial Measurement'!N9",
                40: "'Lease Schedule'!E2", 41: "'Lease Schedule'!I2",
                42: "'Lease Schedule'!K2"}
    for row, ref in expected.items():
        value = ws.cell(row=row, column=2).value
        assert isinstance(value, str) and value.startswith("=") and ref in value


def test_inputs_sheet_keeps_floored_depreciation_periods(calc):
    assert calc["Inputs"]["B33"].value == "=MIN(B14,INT(B32))"


def test_schedule_has_120_row_guarded_periods(calc):
    ws = calc["Lease Schedule"]
    first, last = 4, 4 + kit.MAX_PERIODS - 1
    for row in (first, last):
        for col in (1, 5, 9):                       # period, interest, dep
            value = ws.cell(row=row, column=col).value
            assert isinstance(value, str) and value.startswith("=")
        # every schedule formula self-blanks beyond N
        assert '$A' in ws.cell(row=row, column=9).value
    assert ws.cell(row=last + 1, column=1).value is None


def test_depreciation_formula_divides_by_floored_periods(calc):
    formula = calc["Lease Schedule"].cell(row=4, column=9).value
    assert "Inputs!$B$33" in formula
    assert "'Initial Measurement'!$N$9" in formula


def test_setup_sheet_mirrors_countries_config(calc):
    ws = calc["Setup"]
    rows = {ws.cell(row=r, column=1).value: r for r in range(2, 7)}
    assert set(rows) == set(kit.COUNTRIES)
    for name, r in rows.items():
        c = kit.COUNTRIES[name]
        assert ws.cell(row=r, column=2).value == c["cur"]
        assert ws.cell(row=r, column=5).value == c["std"]
        assert ws.cell(row=r, column=6).value == c["isa540"]
        assert ws.cell(row=r, column=9).value == pytest.approx(c["tax"])


def test_framework_lookups_driven_by_country_cell(calc):
    ws = calc["Inputs"]
    for row in range(49, 54):
        value = ws.cell(row=row, column=2).value
        assert "VLOOKUP($B$48,Setup!" in str(value)


def test_journals_reference_measurement_and_schedule(calc):
    formulas = [str(c.value) for row in calc["Journals"].iter_rows()
                for c in row if isinstance(c.value, str)
                and c.value.startswith("=")]
    joined = " ".join(formulas)
    assert "'Initial Measurement'" in joined
    assert "'Lease Schedule'" in joined


def test_findings_wired_to_annual_summary_checks(calc):
    formulas = [str(c.value) for row in calc["Findings"].iter_rows()
                for c in row if isinstance(c.value, str)
                and c.value.startswith("=")]
    assert any("'Annual Summary'" in f for f in formulas)


def test_audit_procedures_and_pbc_are_fully_populated(calc):
    refs = [calc["Audit_Procedures"].cell(row=r, column=1).value
            for r in range(2, 22)]
    assert refs == [f"P-{i}" for i in range(1, 21)]
    items = [calc["PBC_List"].cell(row=r, column=1).value
             for r in range(3, 21)]
    assert items == list(range(1, 19))


@pytest.mark.parametrize("country", list(kit.COUNTRIES))
def test_workbook_builds_for_every_jurisdiction(tmp_path, country):
    path = str(tmp_path / f"{country[:2]}.xlsx")
    kit.build_calculation_workbook(kit.LeaseInputs(country=country), path)
    assert load_workbook(path).sheetnames == SHEETS

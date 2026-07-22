#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
selftest.py — one-file smoke test for the installed ifrs16kit package.

Run it from Spyder (F5) or a terminal:  python3 selftest.py

No pytest needed. It re-checks the essentials of every layer of the full
suite (tests/ + TESTING.md): the golden benchmark, an independent
closed-form engine, the accounting invariants, the template contract,
the liquid workbook structure, and the demo-isolation control — and it
tells you exactly which ifrs16kit it imported, so a stray ifrs16kit.py
in the working folder can never fool you silently.

Exit code 0 = everything passed.
"""
import contextlib
import io
import os
import sys
import tempfile
from datetime import date

RESULTS = []          # (ok, label, detail)
WARNINGS = []


def check(label, fn):
    try:
        fn()
        RESULTS.append((True, label, ""))
        print(f"  PASS  {label}")
    except AssertionError as e:
        RESULTS.append((False, label, str(e)))
        print(f"  FAIL  {label}\n        → {e}")
    except Exception as e:                              # noqa: BLE001
        RESULTS.append((False, label, f"{type(e).__name__}: {e}"))
        print(f"  FAIL  {label}\n        → {type(e).__name__}: {e}")


def approx(a, b, tol=0.005):
    assert abs(a - b) <= tol, f"{a} != {b} (±{tol})"


# ── 0. Which ifrs16kit did we import? ───────────────────────────────────
import ifrs16kit as kit                                 # noqa: E402

print("═" * 66)
print("  ifrs16kit self-test")
print("═" * 66)
print(f"  imported from : {getattr(kit, '__file__', '?')}")
if hasattr(kit, "__version__"):
    print(f"  version       : {kit.__version__}")
else:
    WARNINGS.append(
        "No __version__ — you imported a standalone ifrs16kit.py from the "
        "current folder, NOT the installed package. Run this from another "
        "directory, or keep the single-file copy in standalone/ instead "
        "of the repo root.")
    print("  version       : (none — see warning at the end)")
print("─" * 66)

from openpyxl import load_workbook                      # noqa: E402


# ── 1. Golden benchmark (CALISMA) ───────────────────────────────────────
def golden():
    cc = kit.cross_check(kit.LeaseInputs())
    approx(cc["liability"], 14_761.09)
    approx(cc["rou"], 15_361.09)
    approx(cc["total_interest"], 838.91)
    approx(cc["total_depreciation"], 15_361.09)
    assert cc["n"] == 24
    approx(cc["closing_liability"], 0, 1e-6)
    approx(cc["closing_rou"], 0, 1e-6)


check("Golden case: liability 14,761.09 · ROU 15,361.09 · interest 838.91",
      golden)


# ── 2. Independent closed-form engine (annuity PV) ──────────────────────
def third_engine():
    for advance in (True, False):
        inp = kit.LeaseInputs(is_advance=advance)
        r, n = inp.periodic_rate, int(inp.num_periods)
        pv = inp.payment * (1 - (1 + r) ** -n) / r
        if advance:
            pv *= (1 + r)
        approx(kit.cross_check(inp)["liability"], pv, 1e-6)


check("Liability matches closed-form annuity (advance and arrears)",
      third_engine)


# ── 3. Accounting invariants across scenarios ───────────────────────────
SCENARIOS = [
    kit.LeaseInputs(),
    kit.LeaseInputs(is_advance=False, freq=4, term_years=3, ibr_annual=0.08),
    kit.LeaseInputs(freq=1, term_years=5, method="Reducing balance",
                    useful_life_months=120),
    kit.LeaseInputs(freq=4, term_years=5, useful_life_months=50),  # fractional
    kit.LeaseInputs(country="Türkiye", freq=2, term_years=4),
]


def invariants():
    for inp in SCENARIOS:
        cc = kit.cross_check(inp)
        tag = f"[{inp.freq}/yr {'adv' if inp.is_advance else 'arr'} {inp.method}]"
        approx(cc["closing_liability"], 0, 1e-6)
        assert abs(cc["liability"] + cc["total_interest"]
                   - cc["total_payments"]) < 1e-6, f"cash identity {tag}"
        assert abs(cc["total_depreciation"] - cc["rou"]) < 1e-6, \
            f"asset not fully depreciated {tag}"
        approx(cc["closing_rou"], 0, 1e-6)


check("Invariants hold across 5 scenarios (incl. fractional useful life)",
      invariants)


# ── 4. Template contract: blank control, validation, round trip ─────────
def template_contract():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "IFRS16_Input.xlsx")
        kit.build_template({"country": "Ireland", "is_advance": True,
                            "freq": 12}, path)
        wb = load_workbook(path)
        ws = wb["Inputs"]
        for cell in ("B5", "B8", "B9", "B15", "B20", "B31"):
            assert ws[cell].value is None, f"{cell} should be blank"
        assert ws["B33"].value == "=MIN(B14,INT(B32))", "B33 must floor B32"

        try:                                   # blank template must reject
            kit.read_template(path)
            raise AssertionError("blank template was accepted")
        except ValueError as e:
            assert "Lessee entity (B5) is empty" in str(e)

        ws["B5"] = "Example Co. Limited"
        ws["B6"] = "Company car"
        ws["B8"] = date(2025, 1, 1)
        ws["B9"] = 2
        ws["B15"] = 650
        ws["B20"] = 0.06
        ws["B31"] = 60
        wb.save(path)
        inp = kit.read_template(path)
        approx(kit.cross_check(inp)["liability"], 14_761.09)


check("Template: blank-required control, cell-level rejection, round trip",
      template_contract)


# ── 5. Workbook: 10 sheets, liquid (no baked numbers) ───────────────────
def workbook_structure():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "calc.xlsx")
        kit.build_calculation_workbook(kit.LeaseInputs(), path)
        wb = load_workbook(path)
        assert len(wb.sheetnames) == 10, wb.sheetnames
        assert wb.sheetnames[0] == "Inputs" and wb.sheetnames[-1] == "Setup"
        for name in ("Initial Measurement", "Lease Schedule",
                     "Annual Summary", "Journals"):
            consts = [c.coordinate for row in wb[name].iter_rows()
                      for c in row if isinstance(c.value, (int, float))
                      and not isinstance(c.value, bool)]
            assert not consts, f"{name} has hard-coded numbers at {consts}"
        b36 = wb["Inputs"]["B36"].value
        assert str(b36).startswith("=") and "Initial Measurement" in b36


check("Workbook: 10 sheets, core sheets fully formula-driven",
      workbook_structure)


# ── 6. Demo isolation: never touches real workbooks ─────────────────────
def demo_isolation():
    with tempfile.TemporaryDirectory() as d:
        sentinel = os.path.join(d, "IFRS16_Calculation.xlsx")
        open(sentinel, "wb").close()
        with contextlib.redirect_stdout(io.StringIO()):   # keep output tidy
            t, o = kit.demo(d)
        assert os.path.basename(t) == "IFRS16_Input_DEMO.xlsx"
        assert os.path.basename(o) == "IFRS16_Calculation_DEMO.xlsx"
        assert os.path.getsize(sentinel) == 0, "demo overwrote a real file"


check("demo() writes only _DEMO files; real workbooks untouched",
      demo_isolation)


# ── Summary ─────────────────────────────────────────────────────────────
failed = [r for r in RESULTS if not r[0]]
print("─" * 66)
print(f"  {len(RESULTS) - len(failed)}/{len(RESULTS)} checks passed")
for warning in WARNINGS:
    print(f"\n  ⚠  {warning}")
if failed:
    print("\n  FAILED:")
    for _, label, detail in failed:
        print(f"   • {label} — {detail}")
print("═" * 66)
sys.exit(1 if failed else 0)

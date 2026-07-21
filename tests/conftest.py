"""Shared fixtures and helpers for the ifrs16kit test suite."""
import os
import shutil
import subprocess
import sys
from datetime import date

import pytest
from openpyxl import load_workbook

import ifrs16kit as kit

# ── The CALISMA golden benchmark (monthly advance, 24 periods, 6% IBR,
#    650/month, 600 IDC) — the values documented in README/VIGNETTE. ──
GOLDEN = {
    "liability": 14_761.09,
    "rou": 15_361.09,
    "total_interest": 838.91,
    "total_depreciation": 15_361.09,
    "total_expense": 16_200.00,
    "total_payments": 15_600.00,
    "n": 24,
}

MONEY = dict(abs=0.005)          # half-a-cent tolerance for money asserts


def fill_template(path, *, entity="Example Co. Limited",
                  asset="Company car (illustrative)",
                  commencement=date(2025, 1, 1), years=2, extra=None,
                  freq=None, payment=650, timing=None, ibr=0.06, idc=None,
                  prepaid=None, incentives=None, restoration=None,
                  method=None, life=60, rb=None, country=None):
    """Fill a generated input template the way a user would in Excel.

    ``None`` means "leave whatever the template already contains" (the
    template pre-writes freq/timing/country from the interview config and
    the non-required example values). Explicit values overwrite cells;
    pass e.g. ``entity=""`` to blank a required cell.
    """
    wb = load_workbook(path)     # keep formulas intact, like Excel would
    ws = wb["Inputs"]
    cells = {"B5": entity, "B6": asset, "B8": commencement, "B9": years,
             "B10": extra, "B12": freq, "B15": payment, "B16": timing,
             "B20": ibr, "B24": idc, "B25": prepaid, "B26": incentives,
             "B27": restoration, "B30": method, "B31": life, "B34": rb,
             "B48": country}
    for cell, value in cells.items():
        if value is not None:
            ws[cell] = value
    wb.save(path)
    return path


@pytest.fixture
def golden_template(tmp_path):
    """A template filled exactly with the golden-case values."""
    path = str(tmp_path / "IFRS16_Input.xlsx")
    kit.build_template({"country": "Ireland", "is_advance": True, "freq": 12},
                       path)
    return fill_template(path)


# ── LibreOffice head-less recalculation (two-engine integration tests) ──

def soffice_cmd():
    """Return the command list to invoke LibreOffice, or None if absent.

    IFRS16KIT_SOFFICE may point to a binary or a .py wrapper (used in
    sandboxed environments); otherwise `soffice` is looked up on PATH.
    """
    override = os.environ.get("IFRS16KIT_SOFFICE")
    if override:
        return ([sys.executable, override] if override.endswith(".py")
                else [override])
    binary = shutil.which("soffice")
    return [binary] if binary else None


def recalc(paths, outdir):
    """Convert workbook(s) via LibreOffice so formula results get cached."""
    cmd = soffice_cmd() + ["--headless", "--convert-to", "xlsx",
                           "--outdir", str(outdir)] + [str(p) for p in paths]
    subprocess.run(cmd, check=True, capture_output=True, timeout=600)
    return [os.path.join(str(outdir), os.path.basename(p)) for p in paths]

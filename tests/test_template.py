"""Layer 2 — the input template contract (build_template / read_template).

The template is a control: required cells arrive blank, everything derived
is a live formula, and read-back validation rejects bad fills cell by cell.
"""
import pytest
from openpyxl import load_workbook, Workbook

import ifrs16kit as kit
from conftest import GOLDEN, MONEY, fill_template

CFG = {"country": "Ireland", "is_advance": True, "freq": 12}
REQUIRED_BLANK = ["B5", "B6", "B8", "B9", "B15", "B20", "B31"]


@pytest.fixture
def blank_template(tmp_path):
    path = str(tmp_path / "IFRS16_Input.xlsx")
    kit.build_template(CFG, path)
    return path


# ── Structure of the generated template ─────────────────────────────────

def test_required_cells_are_blank(blank_template):
    ws = load_workbook(blank_template)["Inputs"]
    for cell in REQUIRED_BLANK:
        assert ws[cell].value is None, f"{cell} should be blank"


def test_interview_config_and_defaults_are_prewritten(blank_template):
    ws = load_workbook(blank_template)["Inputs"]
    assert ws["B12"].value == 12
    assert ws["B16"].value == "Advance"
    assert ws["B48"].value == "Ireland"
    assert ws["B30"].value == "Straight-line"
    assert ws["B10"].value == 0
    assert ws["B24"].value == 600          # example IDC (non-required)


def test_derived_cells_are_live_formulas(blank_template):
    ws = load_workbook(blank_template)["Inputs"]
    assert ws["B11"].value == "=B9*12+B10"
    assert ws["B13"].value == "=12/B12"
    assert ws["B14"].value == "=B11/B13"
    assert ws["B21"].value == "=(1+B20)^(1/B12)-1"
    for cell in ("B44", "B45"):            # green validation checks
        assert str(ws[cell].value).startswith("=IF(")
        assert "OK" in ws[cell].value


def test_depreciation_periods_formula_floors_partial_periods(blank_template):
    # Regression: must floor B32 so Excel matches the Python engine and
    # the reducing-balance final write-off can trigger.
    ws = load_workbook(blank_template)["Inputs"]
    assert ws["B33"].value == "=MIN(B14,INT(B32))"


def test_dropdown_validations_present(blank_template):
    ws = load_workbook(blank_template)["Inputs"]
    assert len(ws.data_validations.dataValidation) == 4


def test_prefill_writes_golden_example():
    import tempfile, os
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "t.xlsx")
        kit.build_template(CFG, path, prefill=True)
        ws = load_workbook(path)["Inputs"]
        assert ws["B5"].value == "Example Co. Limited"
        assert ws["B20"].value == 0.06
        assert ws["B15"].value == 650


# ── Round trip: fill like a user, read back, recompute ──────────────────

def test_round_trip_reproduces_golden_figures(golden_template):
    inp = kit.read_template(golden_template)
    assert inp.entity == "Example Co. Limited"
    assert inp.country == "Ireland"
    assert (inp.term_months, inp.freq, int(inp.num_periods)) == (24, 12, 24)
    assert inp.is_advance and inp.method == "Straight-line"
    assert (inp.payment, inp.ibr_annual, inp.idc) == (650, 0.06, 600)
    cc = kit.cross_check(inp)
    assert cc["liability"] == pytest.approx(GOLDEN["liability"], **MONEY)
    assert cc["rou"] == pytest.approx(GOLDEN["rou"], **MONEY)


def test_percent_formatted_ibr_reads_as_decimal(blank_template):
    # Typing "6%" in Excel stores 0.06 — the template read must accept it.
    fill_template(blank_template, ibr=0.06)
    assert kit.read_template(blank_template).ibr_annual == pytest.approx(0.06)


@pytest.mark.parametrize("raw,expected", [
    ("arrears", False), ("ADVANCE", True), ("Advance", True)])
def test_timing_case_insensitive(blank_template, raw, expected):
    fill_template(blank_template, timing=raw)
    assert kit.read_template(blank_template).is_advance is expected


@pytest.mark.parametrize("raw,norm", [
    ("ireland", "Ireland"), ("TÜRKIYE", "Türkiye"), ("uae", "UAE")])
def test_country_case_normalisation(blank_template, raw, norm):
    fill_template(blank_template, country=raw)
    assert kit.read_template(blank_template).country == norm


def test_method_case_normalisation(blank_template):
    fill_template(blank_template, method="reducing balance")
    assert kit.read_template(blank_template).method == "Reducing balance"


def test_iso_date_string_accepted(blank_template):
    fill_template(blank_template, commencement="2025-03-01")
    assert str(kit.read_template(blank_template).commencement) == "2025-03-01"


# ── Cell-by-cell rejection (the ISA 230 traceability property) ──────────

def _err(path):
    with pytest.raises(ValueError) as e:
        kit.read_template(path)
    return str(e.value)


def test_unfilled_template_lists_every_required_cell(blank_template):
    msg = _err(blank_template)
    for fragment in ["Lessee entity (B5) is empty",
                     "Commencement date (B8) is empty",
                     "Lease term — YEARS (B9) is empty",
                     "Lease payment (B15) is empty",
                     "Annual IBR (B20) is empty",
                     "Useful life (B31) is empty"]:
        assert fragment in msg


def test_ibr_entered_as_whole_percentage(blank_template):
    fill_template(blank_template, ibr=6)
    msg = _err(blank_template)
    assert "looks like a percentage" in msg
    # Regression: previously duplicated with a second IBR error line.
    assert "must be between 0 and 1" not in msg


def test_ibr_zero_rejected(blank_template):
    fill_template(blank_template, ibr=0)
    assert "must be between 0 and 1" in _err(blank_template)


def test_non_whole_number_of_periods_rejected(blank_template):
    fill_template(blank_template, years=2, extra=1, freq=4)   # 25 months / Q
    assert "whole number" in _err(blank_template)


def test_capacity_of_120_periods_enforced(blank_template):
    fill_template(blank_template, years=11, freq=12)          # N = 132
    assert f"template capacity" in _err(blank_template)
    assert "132" in _err(blank_template)


def test_timing_typo_rejected_not_silently_arrears(blank_template):
    # Regression: free-text typos used to fall through to Arrears.
    fill_template(blank_template, timing="Advnace")
    assert "must be Advance or Arrears" in _err(blank_template)


@pytest.mark.parametrize("kw,fragment", [
    (dict(payment=-5), "Lease payment (B15) must be > 0"),
    (dict(payment="abc"), "not a number"),
    (dict(years=0), "Lease term must be > 0"),
    (dict(life=0), "Useful life (B31) must be > 0"),
    (dict(country="France"), "Country (B48) must be one of"),
    (dict(freq=5), "Frequency (B12) must be 12/4/2/1"),
    (dict(method="Units of production"), "must be Straight-line or Reducing balance"),
    (dict(commencement="soon"), "must be a date"),
])
def test_individual_validation_rules(blank_template, kw, fragment):
    fill_template(blank_template, **kw)
    assert fragment in _err(blank_template)


def test_wrong_workbook_without_inputs_sheet(tmp_path):
    p = str(tmp_path / "other.xlsx")
    Workbook().save(p)
    with pytest.raises(ValueError, match="Sheet 'Inputs' not found"):
        kit.read_template(p)

"""Layer 6 — the two-engine control, fully automated.

Builds liquid workbooks, has LibreOffice recalculate them head-lessly, and
compares Excel's computed figures against the independent Python engine
(`cross_check`). This automates the ISA 500 accuracy-and-completeness
verification of the auditor's own tool that the vignette performs manually.

Skipped automatically when LibreOffice is unavailable; CI installs it.
"""
import pytest
from openpyxl import load_workbook

import ifrs16kit as kit
from conftest import soffice_cmd, recalc

pytestmark = pytest.mark.skipif(soffice_cmd() is None,
                                reason="LibreOffice not available")

SCENARIOS = {
    "golden_monthly_advance": kit.LeaseInputs(),
    "arrears_quarterly": kit.LeaseInputs(is_advance=False, freq=4,
                                         term_years=3, ibr_annual=0.08),
    "annual_uk_reducing": kit.LeaseInputs(country="United Kingdom", freq=1,
                                          term_years=5, ibr_annual=0.07,
                                          method="Reducing balance",
                                          useful_life_months=120),
    # Regression: fractional useful-life-in-periods, life-limited.
    "fractional_life_regression": kit.LeaseInputs(freq=4, term_years=5,
                                                  useful_life_months=50),
}

# Inputs sheet key-output rows → cross_check keys.
ROWS = {36: "liability", 37: "rou", 40: "total_interest",
        41: "total_depreciation"}


@pytest.fixture(scope="module")
def recalculated(tmp_path_factory):
    src = tmp_path_factory.mktemp("src")
    out = tmp_path_factory.mktemp("recalced")
    paths = []
    for name, inp in SCENARIOS.items():
        p = src / f"{name}.xlsx"
        kit.build_calculation_workbook(inp, str(p))
        paths.append(p)
    converted = recalc(paths, out)
    return dict(zip(SCENARIOS, converted))


@pytest.mark.parametrize("name", list(SCENARIOS))
def test_excel_engine_matches_python_engine(recalculated, name):
    cc = kit.cross_check(SCENARIOS[name])
    ws = load_workbook(recalculated[name], data_only=True)["Inputs"]
    for row, key in ROWS.items():
        excel = ws.cell(row=row, column=2).value
        assert excel == pytest.approx(cc[key], abs=0.005), (
            f"{name}: Inputs!B{row} ({key}) Excel={excel} Python={cc[key]}")
    # Total expense (B42) = interest + depreciation.
    expense = ws.cell(row=42, column=2).value
    assert expense == pytest.approx(
        cc["total_interest"] + cc["total_depreciation"], abs=0.005)


def test_annual_summary_reconciliation_checks_all_ok(recalculated):
    ws = load_workbook(recalculated["golden_monthly_advance"],
                       data_only=True)["Annual Summary"]
    statuses = [c.value for row in ws.iter_rows() for c in row
                if c.value in ("OK",) or (isinstance(c.value, str)
                                          and c.value.startswith("CHECK"))]
    ok = [s for s in statuses if s == "OK"]
    failed = [s for s in statuses if s != "OK"]
    assert len(ok) >= 5 and not failed, f"failed checks: {failed}"


def test_fractional_life_fully_depreciates_in_excel(recalculated):
    """The Excel-side proof of the INT(B32) fix: total depreciation equals
    the ROU asset even when useful life is not a whole number of periods."""
    ws = load_workbook(recalculated["fractional_life_regression"],
                       data_only=True)["Inputs"]
    rou = ws.cell(row=37, column=2).value
    total_dep = ws.cell(row=41, column=2).value
    assert total_dep == pytest.approx(rou, abs=0.005)

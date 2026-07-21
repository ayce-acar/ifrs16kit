"""Layer 5 — the public API surface and packaging metadata."""
import importlib.metadata

import pytest

import ifrs16kit as kit
from conftest import fill_template


def test_all_exports_resolve():
    for name in kit.__all__:
        assert getattr(kit, name) is not None, name


def test_version_single_sourced_with_package_metadata():
    assert kit.__version__ == importlib.metadata.version("ifrs16kit")


def test_constants():
    assert kit.MAX_PERIODS == 120
    assert set(kit.FREQUENCIES) == {12, 4, 2, 1}
    assert set(kit.COUNTRIES) == {"Türkiye", "Ireland", "United Kingdom",
                                  "UAE", "Australia"}
    for cfg in kit.COUNTRIES.values():
        assert {"cur", "sym", "framework", "std",
                "isa540", "isa500", "isa230", "tax"} <= set(cfg)


def test_create_input_writes_template_and_returns_path(tmp_path):
    path = kit.create_input(str(tmp_path / "in.xlsx"), country="Türkiye",
                            advance=False, freq=4)
    assert path.endswith("in.xlsx")
    from openpyxl import load_workbook
    ws = load_workbook(path)["Inputs"]
    assert ws["B48"].value == "Türkiye"
    assert ws["B16"].value == "Arrears"
    assert ws["B12"].value == 4


@pytest.mark.parametrize("kw", [dict(country="France"), dict(freq=7)])
def test_create_input_rejects_bad_arguments(tmp_path, kw):
    with pytest.raises(ValueError):
        kit.create_input(str(tmp_path / "x.xlsx"), **kw)


def test_create_calculation_from_filled_template(tmp_path, capsys):
    template = str(tmp_path / "in.xlsx")
    output = str(tmp_path / "out.xlsx")
    kit.create_input(template)
    fill_template(template)
    result = kit.create_calculation(template, output)
    assert result.endswith("out.xlsx")
    out = capsys.readouterr().out
    assert "Inputs read from the template" in out
    from openpyxl import load_workbook
    assert len(load_workbook(output).sheetnames) == 10


def test_create_calculation_propagates_validation_errors(tmp_path):
    template = str(tmp_path / "in.xlsx")
    kit.create_input(template)                 # left blank on purpose
    with pytest.raises(ValueError, match="is empty"):
        kit.create_calculation(template, str(tmp_path / "out.xlsx"))

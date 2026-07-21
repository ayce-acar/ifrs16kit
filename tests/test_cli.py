"""Layer 4 — the CLI and the interactive flow.

Covers `--demo` (including the never-overwrite regression), `python -m
ifrs16kit`, `demo()`, and the full scripted interview → fill → validate →
build loop, including the fix-and-retry path.
"""
import os
import subprocess
import sys

import pytest
from openpyxl import load_workbook

import ifrs16kit as kit
from ifrs16kit import core
from conftest import fill_template


# ── demo() helper ───────────────────────────────────────────────────────

def test_demo_writes_both_demo_files(tmp_path):
    t, o = kit.demo(str(tmp_path))
    assert os.path.basename(t) == "IFRS16_Input_DEMO.xlsx"
    assert os.path.basename(o) == "IFRS16_Calculation_DEMO.xlsx"
    assert os.path.exists(t) and os.path.exists(o)
    assert load_workbook(o).sheetnames[0] == "Inputs"


# ── --demo via main(): the overwrite regression ─────────────────────────

def test_cli_demo_never_touches_real_workbooks(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    sentinel_calc = tmp_path / "IFRS16_Calculation.xlsx"
    sentinel_tmpl = tmp_path / "IFRS16_Input.xlsx"
    sentinel_calc.write_bytes(b"")            # pretend real files exist
    sentinel_tmpl.write_bytes(b"")
    monkeypatch.setattr(sys, "argv", ["ifrs16kit", "--demo"])
    core.main()
    out = capsys.readouterr().out
    assert "14,761.09" in out and "15,361.09" in out
    assert sentinel_calc.read_bytes() == b""  # untouched
    assert sentinel_tmpl.read_bytes() == b""
    assert (tmp_path / "IFRS16_Calculation_DEMO.xlsx").exists()
    assert (tmp_path / "IFRS16_Input_DEMO.xlsx").exists()


def test_python_dash_m_entry_point(tmp_path):
    result = subprocess.run(
        [sys.executable, "-m", "ifrs16kit", "--demo"],
        cwd=tmp_path, capture_output=True, text=True, timeout=120)
    assert result.returncode == 0
    assert "14,761.09" in result.stdout
    assert (tmp_path / "IFRS16_Calculation_DEMO.xlsx").exists()


# ── Full interactive flow, scripted ─────────────────────────────────────

class ScriptedInput:
    """Feeds answers to input(); fills the template at the pause prompt."""

    def __init__(self, template_path, answers=None, fill_kwargs=None):
        self.template = str(template_path)
        self.answers = list(answers or [])
        self.fill_kwargs = fill_kwargs or {}
        self.filled = False
        self.prompts = []

    def __call__(self, prompt=""):
        self.prompts.append(prompt)
        if "press enter" in prompt.lower() and not self.filled:
            fill_template(self.template, **self.fill_kwargs)
            self.filled = True
            return ""
        return self.answers.pop(0) if self.answers else ""


def test_interactive_flow_end_to_end(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    template = tmp_path / core.TEMPLATE_NAME
    script = ScriptedInput(template)          # defaults everywhere
    monkeypatch.setattr("builtins.input", script)
    monkeypatch.setattr(sys, "argv", ["ifrs16kit"])
    core.run()                                # run is main
    out = capsys.readouterr().out
    assert "Template created" in out
    assert "Extracted inputs" in out
    assert (tmp_path / core.OUTPUT_NAME).exists()
    assert load_workbook(tmp_path / core.OUTPUT_NAME).sheetnames[0] == "Inputs"


def test_interactive_flow_retries_until_template_valid(tmp_path, monkeypatch,
                                                       capsys):
    monkeypatch.chdir(tmp_path)
    template = tmp_path / core.TEMPLATE_NAME

    class BadThenGood(ScriptedInput):
        def __call__(self, prompt=""):
            self.prompts.append(prompt)
            if "press enter" in prompt.lower():
                if not self.filled:           # first pass: invalid IBR
                    fill_template(self.template, ibr=6)
                    self.filled = True
                else:                          # second pass: fix it
                    fill_template(self.template, ibr=0.06)
                return ""
            return ""

    monkeypatch.setattr("builtins.input", BadThenGood(template))
    core.main()
    out = capsys.readouterr().out
    assert "looks like a percentage" in out    # the rejection was surfaced
    assert (tmp_path / core.OUTPUT_NAME).exists()


def test_existing_filled_template_is_reused(tmp_path, monkeypatch, capsys):
    monkeypatch.chdir(tmp_path)
    template = str(tmp_path / core.TEMPLATE_NAME)
    kit.build_template({"country": "Ireland", "is_advance": True, "freq": 12},
                       template)
    fill_template(template, entity="Reuse Co. Limited")
    monkeypatch.setattr("builtins.input", ScriptedInput(template))
    core.main()                                # "" → reuse, "" → confirm
    out = capsys.readouterr().out
    assert "Found existing template" in out
    assert "Reuse Co. Limited" in out
    assert (tmp_path / core.OUTPUT_NAME).exists()


# ── ask_choice / ask_yes_no ─────────────────────────────────────────────

def test_ask_choice_validates_input(monkeypatch, capsys):
    feed = iter(["7", "x", "2"])
    monkeypatch.setattr("builtins.input", lambda _: next(feed))
    assert core.ask_choice("Pick", ["a", "b", "c"], 0) == 1
    assert "Invalid" in capsys.readouterr().out


def test_ask_yes_no_defaults_and_parses(monkeypatch):
    monkeypatch.setattr("builtins.input", lambda _: "")
    assert core.ask_yes_no("Q", True) is True
    monkeypatch.setattr("builtins.input", lambda _: "nope")
    assert core.ask_yes_no("Q", True) is False


def test_print_summary_shows_golden_figures(capsys):
    kit.print_summary(kit.LeaseInputs())
    out = capsys.readouterr().out
    assert "14,761.09" in out and "RECONCILED" in out
